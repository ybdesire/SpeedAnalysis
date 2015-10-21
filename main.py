import timeit
from DataCalc import DataCalc
from CalcDistance import calc_distance
from openpyxl import Workbook

def test():
	distance = calc_distance(121.306216, 31.353354, 121.185040, 31.405485)
	print(distance)#km
	
	
FILENAME = 'data-20150917.xlsx'

def main():
	data_calc_mgr = DataCalc(FILENAME, 'vehicle_gps_log')
	speed_zero_index_list = data_calc_mgr.get_car_stop_point()	
	print('speed zero point\n{0}'.format(speed_zero_index_list))

	car_stop_area = data_calc_mgr.get_car_stop_area(speed_zero_index_list)
	print('stop area\n{0}'.format(car_stop_area))
	
	valid_stop_area = data_calc_mgr.get_valid_stop_area(car_stop_area)
	print('valid stop area\n{0}'.format(valid_stop_area))

	stop_lon_lat = data_calc_mgr.get_stop_lon_lat(valid_stop_area)
	print('\ncar stop lon lat:')
	print_order_dict(stop_lon_lat)
	
	car_stop_start_time = data_calc_mgr.get_stop_start_time(valid_stop_area)
	print('\ncar stop start time:')
	print_order_dict(car_stop_start_time)
	
	car_stop_end_time  = data_calc_mgr.get_stop_end_time(valid_stop_area)
	print('\ncar stop end time:')
	print_order_dict(car_stop_end_time)
	
	car_stop_time_interval = data_calc_mgr.get_stop_time_interval(valid_stop_area)
	print('\ncar stop time intervel:')
	print_order_dict(car_stop_time_interval)
	
	car_stop_time_sum = data_calc_mgr.get_stop_time_sum(car_stop_time_interval)
	print('\ncar stop time sum:')
	print(car_stop_time_sum)
	
	car_driving_area = data_calc_mgr.get_driving_area(valid_stop_area)
	print('\ncar driving area:')
	print_order_dict(car_driving_area)
	
	driving_area_ave_speed = data_calc_mgr.get_driving_ave_speed(car_driving_area)
	print('\ncar driving average speed:')
	print_order_dict(driving_area_ave_speed)	
	
	car_driving_time_stat = data_calc_mgr.get_driving_time_stat(car_driving_area)
	print('\ncar driving start-time, end-time, intervel:')
	print_order_dict(car_driving_time_stat)	
	
	car_driving_area_time_interval = data_calc_mgr.get_driving_time_interval(car_driving_area)
	car_driving_time_sum = data_calc_mgr.get_driving_time_sum(car_driving_area_time_interval)
	print('\ncar driving time sum:')
	print(car_driving_time_sum)
	
	car_driving_lon_lat = data_calc_mgr.get_driving_lon_lat(car_driving_area)
	print('\ncar driving start-lon lat, end-lon lat')
	print_order_dict(car_driving_lon_lat)	
	
	car_driving_length = data_calc_mgr.get_driving_path_length_list(car_driving_area)
	print('\ncar driving path length:')
	print_order_dict(car_driving_length)
	
	car_driving_path_sum = data_calc_mgr.get_driving_length_sum(car_driving_length)
	print('\ncar driving path length sum:')
	print(car_driving_path_sum)
	
	speed_variance_70 = data_calc_mgr.get_speed_variance(car_driving_area, 70)
	print('\ncar speed variance with 70km/h')
	print_order_dict(speed_variance_70)	

	speed_variance_75 = data_calc_mgr.get_speed_variance(car_driving_area, 75)
	print('\ncar speed variance with 75km/h')
	print_order_dict(speed_variance_75)	

	speed_variance_80 = data_calc_mgr.get_speed_variance(car_driving_area, 80)
	print('\ncar speed variance with 80km/h')
	print_order_dict(speed_variance_80)	

	speed_variance_85 = data_calc_mgr.get_speed_variance(car_driving_area, 85)
	print('\ncar speed variance with 85km/h')
	print_order_dict(speed_variance_85)	

	speed_variance_90 = data_calc_mgr.get_speed_variance(car_driving_area, 90)
	print('\ncar speed variance with 90km/h')
	print_order_dict(speed_variance_90)	

	speed_variance_95 = data_calc_mgr.get_speed_variance(car_driving_area, 95)
	print('\ncar speed variance with 95km/h')
	print_order_dict(speed_variance_95)	

	
	car_driving_reasonable_speed_time_interval_77_83 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 77, 83)
	car_driving_reasonable_speed_time_interval_0_20 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 0, 20)
	car_driving_reasonable_speed_time_interval_20_40 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 20, 40)
	car_driving_reasonable_speed_time_interval_40_60 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 40, 60)
	car_driving_reasonable_speed_time_interval_60_75 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 60, 75)
	car_driving_reasonable_speed_time_interval_75_85 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 75, 85)
	car_driving_reasonable_speed_time_interval_85_95 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 85, 95)
	car_driving_reasonable_speed_time_interval_95_1000000 = data_calc_mgr.get_reasonable_speed_time_interval(car_driving_area, 95, 100000)
	print('\ncar driving reasonable speed tiem interval')
	print('speed {0}-{1}: {2} min'.format(0, 20, car_driving_reasonable_speed_time_interval_0_20))	
	print('speed {0}-{1}: {2} min'.format(20, 40, car_driving_reasonable_speed_time_interval_20_40))	
	print('speed {0}-{1}: {2} min'.format(40, 60, car_driving_reasonable_speed_time_interval_40_60))	
	print('speed {0}-{1}: {2} min'.format(60, 75, car_driving_reasonable_speed_time_interval_60_75))	
	print('speed {0}-{1}: {2} min'.format(75, 85, car_driving_reasonable_speed_time_interval_75_85))	
	print('speed {0}-{1}: {2} min'.format(85, 95, car_driving_reasonable_speed_time_interval_85_95))	
	print('speed {0}-{1}: {2} min'.format(95, 1000000, car_driving_reasonable_speed_time_interval_95_1000000))
		
	acc_area, acc_index_area = data_calc_mgr.get_speed_change_area(car_driving_area)
	print('\nacc area')
	print(acc_area)
	acc_area_speed_inc, index_area_speed_inc = data_calc_mgr.get_speed_increase_area(acc_area, acc_index_area)
	speed_inc_start_time = data_calc_mgr.get_speed_inc_dec_start_time(index_area_speed_inc)
	speed_inc_end_time = data_calc_mgr.get_speed_inc_dec_end_time(index_area_speed_inc)
	speed_inc_time_interval = data_calc_mgr.get_speed_inc_dec_time_interval(index_area_speed_inc)
	speed_inc_speed_vari = data_calc_mgr.get_speed_inc_dec_speed(index_area_speed_inc)
	speed_inc_acc = data_calc_mgr.get_speed_inc_dec_acc(index_area_speed_inc)
	speed_inc_acc_slow_take_off = data_calc_mgr.get_slow_take_off_from_inc_acc(index_area_speed_inc, 0.15)
	speed_inc_acc_joggling = data_calc_mgr.get_joggling_acc(index_area_speed_inc)
	
	speed_inc_wb = Workbook()
	speed_inc_ws = speed_inc_wb.active
	
	speed_inc_ws['A1'] = 'acc_list'
	print_to_excel(speed_inc_ws, 'A', acc_area_speed_inc)
	speed_inc_ws['B1'] = 'index_list'
	print_to_excel(speed_inc_ws, 'B', index_area_speed_inc)
	speed_inc_ws['C1'] = 'start_time'
	print_to_excel(speed_inc_ws, 'C', speed_inc_start_time)
	speed_inc_ws['D1'] = 'end_time'
	print_to_excel(speed_inc_ws, 'D', speed_inc_end_time)
	speed_inc_ws['E1'] = 'time_interval'
	print_to_excel(speed_inc_ws, 'E', speed_inc_time_interval)
	speed_inc_ws['F1'] = 'speed_vari'
	print_to_excel(speed_inc_ws, 'F', speed_inc_speed_vari)
	speed_inc_ws['G1'] = 'ave_acc'
	print_to_excel(speed_inc_ws, 'G', speed_inc_acc)
	speed_inc_ws['H1'] = 'slow_take_off'
	print_to_excel(speed_inc_ws, 'H', speed_inc_acc_slow_take_off)
	speed_inc_ws['I1'] = 'speed_joggling'
	print_to_excel(speed_inc_ws, 'I', speed_inc_acc_joggling)
	
	speed_inc_wb.save("speed_inc_state.xlsx")
	print('\nspeed increase stat calculation ended.')
	
	acc_area_speed_dec, index_area_speed_dec = data_calc_mgr.get_speed_decrease_area(acc_area, acc_index_area)
	speed_dec_start_time = data_calc_mgr.get_speed_inc_dec_start_time(index_area_speed_dec)
	speed_dec_end_time = data_calc_mgr.get_speed_inc_dec_end_time(index_area_speed_dec)
	speed_dec_time_interval = data_calc_mgr.get_speed_inc_dec_time_interval(index_area_speed_dec)
	speed_dec_speed_vari = data_calc_mgr.get_speed_inc_dec_speed(index_area_speed_dec)
	speed_dec_acc = data_calc_mgr.get_speed_inc_dec_acc(index_area_speed_dec)
	speed_dec_acc_quick_slow_down = data_calc_mgr.get_quick_slow_down_from_dec_acc(index_area_speed_dec, -0.6)
	speed_dec_acc_joggling = data_calc_mgr.get_joggling_acc(index_area_speed_dec)

	speed_dec_wb = Workbook()
	speed_dec_ws = speed_dec_wb.active
	
	speed_dec_ws['A1'] = 'acc_list'
	print_to_excel(speed_dec_ws, 'A', acc_area_speed_dec)
	speed_dec_ws['B1'] = 'index_list'
	print_to_excel(speed_dec_ws, 'B', index_area_speed_dec)
	speed_dec_ws['C1'] = 'start_time'
	print_to_excel(speed_dec_ws, 'C', speed_dec_start_time)
	speed_dec_ws['D1'] = 'end_time'
	print_to_excel(speed_dec_ws, 'D', speed_dec_end_time)
	speed_dec_ws['E1'] = 'time_interval'
	print_to_excel(speed_dec_ws, 'E', speed_dec_time_interval)
	speed_dec_ws['F1'] = 'speed_vari'
	print_to_excel(speed_dec_ws, 'F', speed_dec_speed_vari)
	speed_dec_ws['G1'] = 'ave_acc'
	print_to_excel(speed_dec_ws, 'G', speed_dec_acc)
	speed_dec_ws['H1'] = 'quick_slow_down'
	print_to_excel(speed_dec_ws, 'H', speed_dec_acc_quick_slow_down)
	speed_dec_ws['I1'] = 'speed_joggling'
	print_to_excel(speed_dec_ws, 'I', speed_dec_acc_joggling)
	
	speed_dec_wb.save("speed_dec_state.xlsx")
	print('\nspeed decrease stat calculation ended.')

	
def print_order_dict(data):
	for i in range(len(data)):
		print(data[i])

		
def print_to_excel(ws, col, data_dict):
	for i in range(len(data_dict)):
		print(data_dict[i])
		if(type(data_dict[i])==type([1,2,3])):
			ws[col+str(i+2)] = ','.join(str(x) for x in data_dict[i])
		else:
			ws[col+str(i+2)] = data_dict[i]
	
	
if __name__=='__main__':
	start = timeit.default_timer()
	main()
	stop = timeit.default_timer()
	print('------ running time: {0} s ------'.format(stop-start))	
	
