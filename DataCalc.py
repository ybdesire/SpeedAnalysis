from CalcDistance import calc_distance
from openpyxl import load_workbook, Workbook
from datetime import datetime
import math

AREA_DIST_THRESHOLD = 0.3 # 0.3KM
ZERO_SPEED_THRESHOLD = 10 #10KM/H. car stop speed cannot be real 0.
DIST_THRESHOLD = 0.3 # 0.3KM
ACCE_ZERO_NEG = -0.2 # m/s^2
ACCE_ZERO_POS = 0.2 # m/s^2

class DataCalc:
	def __init__(self, file_path, sheet_name='vehicle_gps_log'):
		self.wb = load_workbook(file_path)
		self.ws = self.wb[sheet_name]
		self.column_length = 0
		
	def __log_to_file(self, filepath, data_dict):
		file = open(filepath, 'w+')
		for i in range(len(data_dict)):
			file.write('{0}:{1}\n'.format(i, data_dict[i]))
		
	def __get_column_length(self):	
		if(self.column_length == 0):
			self.column_length = self.ws.get_highest_row()
		return self.column_length
								
	def get_car_stop_point(self):#return the row number for those whose speed < ZERO_SPEED_THRESHOLD
		self.__get_column_length()
		speed_zero_index_list = []
		for i in range(self.column_length):
			if(i>2):
				current_gps_stat = self.ws['H' + str(i)].value
				pre_gps_stat = self.ws['H' + str(i-1)].value
				current_time = datetime.strptime(str(self.ws['I' + str(i)].value), '%Y-%m-%d %H:%M:%S')
				if(current_gps_stat!='V' and pre_gps_stat!='V' and current_time.year!=1999):
					current_speed = int(self.ws['F' + str(i)].value)
					if(current_speed<=ZERO_SPEED_THRESHOLD):
						speed_zero_index_list.append(i)
			elif(i==2):
				current_gps_stat = self.ws['H' + str(i)].value
				current_time = datetime.strptime(str(self.ws['I' + str(i)].value), '%Y-%m-%d %H:%M:%S')
				if(current_gps_stat!='V' and current_time.year!=1999):
					current_speed = int(self.ws['F' + str(i)].value)
					if(current_speed<=ZERO_SPEED_THRESHOLD):
						speed_zero_index_list.append(i)
		return speed_zero_index_list
		
	def get_car_stop_area(self, speed_zero_index_list):#return the car stop area row lists
		icount = 0
		car_stop_area = {}
		car_stop_list = []
		for i in range(len(speed_zero_index_list)):
			if(i<len(speed_zero_index_list)-1):
				lon1 = float(self.ws['D' + str(speed_zero_index_list[i])].value)
				lat1 = float(self.ws['E' + str(speed_zero_index_list[i])].value)
				lon2 = float(self.ws['D' + str(speed_zero_index_list[i+1])].value)
				lat2 = float(self.ws['E' + str(speed_zero_index_list[i+1])].value)
				dist = calc_distance(lon1, lat1, lon2, lat2)
				if(dist<AREA_DIST_THRESHOLD):
					car_stop_list.append(speed_zero_index_list[i])
				else:
					car_stop_list.append(speed_zero_index_list[i])
					car_stop_area[icount] = car_stop_list
					car_stop_list = []
					icount = icount + 1
		return car_stop_area
		
	def __is_valid_dist_interval(self, index_list):#if max distance bigger than 3km
		max_dist = 0
		for i in index_list:
			for j in index_list:
				lon1 = float(self.ws['D' + str(i)].value)
				lat1 = float(self.ws['E' + str(i)].value)
				lon2 = float(self.ws['D' + str(j)].value)
				lat2 = float(self.ws['E' + str(j)].value)
				dist = calc_distance(lon1, lat1, lon2, lat2)
				if(dist>max_dist):
					max_dist = dist
		if(max_dist>DIST_THRESHOLD):
			return False
		
		return True
		
	def get_valid_stop_area(self, car_stop_area):
		len_area_temp = len(car_stop_area)
		valid_stop_area_temp = car_stop_area
		invalid_key_list = []
		for key in valid_stop_area_temp:
			value = valid_stop_area_temp[key]
			zero_speed_exist_count = 0
			for value_item in value:
				speed = int(self.ws['F' + str(value_item)].value)
				if(speed==0):
					zero_speed_exist_count = zero_speed_exist_count+1
			if (not zero_speed_exist_count>=2 or len(value)==1 or not self.__is_valid_dist_interval(value)):
				invalid_key_list.append(key)
		#remove invalid stop area		
		for k in invalid_key_list:
			del(valid_stop_area_temp[k])
		#sort the key from 1 to n
		valid_stop_area = {}
		j = 0
		for i in range(len_area_temp):
			if i in valid_stop_area_temp:
				valid_stop_area[j] = valid_stop_area_temp[i]
				j = j+1
			
		return 	valid_stop_area

	def get_stop_lon_lat(self, valid_stop_area):
		stop_lon_lat = {}
		for i in range(len(valid_stop_area)):
			stop_index = valid_stop_area[i][0]
			lon = float(self.ws['D' + str(stop_index)].value)
			lat = float(self.ws['E' + str(stop_index)].value)
			stop_lon_lat[i] = [lon, lat]
		return stop_lon_lat

	def get_stop_start_time(self, valid_stop_area):
		stop_time = {}
		for i in range(len(valid_stop_area)):
			stop_index = valid_stop_area[i][0]
			str_time = str(self.ws['I' + str(stop_index)].value)
			stop_time[i] = str_time
			
		return stop_time
		
	def get_stop_end_time(self, valid_stop_area):
		stop_time = {}
		for i in range(len(valid_stop_area)):
			length = len(valid_stop_area[i])
			stop_index = valid_stop_area[i][length-1]
			str_time = str(self.ws['I' + str(stop_index+1)].value)
			stop_time[i] = str_time
			
		return stop_time
	
	def get_stop_time_interval(self, valid_stop_area):
		stop_interval_dict = {}
		for i in range(len(valid_stop_area)):
			length = len(valid_stop_area[i])
			start_index = valid_stop_area[i][0]
			end_index = valid_stop_area[i][length-1]
			start_time = datetime.strptime(str(self.ws['I' + str(start_index)].value), '%Y-%m-%d %H:%M:%S')
			end_time = datetime.strptime(str(self.ws['I' + str(end_index+1)].value), '%Y-%m-%d %H:%M:%S')
			stop_interval_dict[i] = (end_time-start_time).seconds/60
		return stop_interval_dict
	
	def get_stop_time_sum(self, car_stop_time_interval):
		time_sum = 0
		for i in range(len(car_stop_time_interval)):
			time_sum = time_sum+car_stop_time_interval[i]
		return time_sum
	
	def get_driving_area(self, valid_stop_area):
		self.__get_column_length()
		driving_area = {}
		j = 0
		for i in range(len(valid_stop_area)+1):
			if(i==0):
				car_stop_start_index = valid_stop_area[i][0]
				if(car_stop_start_index>2):
					car_driving_start_index = 2
					car_driving_end_index = car_stop_start_index - 1
				else:
					continue
			elif(i==len(valid_stop_area)):
				length_pre = len(valid_stop_area[i-1])
				car_stop_end_index = valid_stop_area[i-1][length_pre-1]
				if(car_stop_end_index<self.column_length-1):
					car_driving_start_index = car_stop_end_index + 1
				else:
					car_driving_start_index = car_stop_end_index
				car_driving_end_index = self.column_length
			else:
				length_pre = len(valid_stop_area[i-1])
				car_driving_start_index = valid_stop_area[i-1][length_pre-1] + 1
				car_driving_end_index = valid_stop_area[i][0] - 1
			driving_area[j] = [car_driving_start_index, car_driving_end_index]
			j = j+1
		return driving_area
	
	def __is_valid_data(self, index):
		gps_stat = self.ws['H' + str(index)].value
		data_time = datetime.strptime(str(self.ws['I' + str(index)].value), '%Y-%m-%d %H:%M:%S')
		if(index>1 and gps_stat!='V' and data_time.year!=1999):
			return True
		else:
			return False
	
	def get_driving_ave_speed(self, driving_area):
		driving_area_ave_speed = {}
		for i in range(len(driving_area)):
			j = driving_area[i][0]
			path_sum = 0
			time_sum = 0
			while(j<=driving_area[i][1]):
				if(j<self.column_length and self.__is_valid_data(j) and self.__is_valid_data(j+1)):
					lon_start = float(self.ws['D' + str(j)].value)
					lat_start = float(self.ws['E' + str(j)].value)
					lon_end = float(self.ws['D' + str(j+1)].value)
					lat_end = float(self.ws['E' + str(j+1)].value)
					path = calc_distance(lon_start, lat_start, lon_end, lat_end)
					path_sum = path_sum + path
					
					start_time = str(self.ws['I' + str(j)].value)			
					end_time = str(self.ws['I' + str(j+1)].value)

					interval = (datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')-datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')).seconds/(60*60)
					time_sum = time_sum + interval
				j = j+1
			driving_area_ave_speed[i] = path_sum/time_sum
		return driving_area_ave_speed
	
	def get_driving_time_stat(self, driving_area):
		driving_area_time = {}
		for i in range(len(driving_area)):
			length = len(driving_area[i])
			start_index = driving_area[i][0]
			end_index = driving_area[i][length-1]
			start_time = str(self.ws['I' + str(start_index)].value)		
			if(end_index!=self.column_length):
				end_index = end_index + 1
			end_time = str(self.ws['I' + str(end_index)].value)
			interval = (datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')-datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')).seconds/60
			driving_area_time[i] = [start_time, end_time, interval]
		return driving_area_time

	def get_driving_time_interval(self, driving_area):
		driving_area_time_interval = {}
		for i in range(len(driving_area)):
			length = len(driving_area[i])
			start_index = driving_area[i][0]
			end_index = driving_area[i][length-1]
			start_time = str(self.ws['I' + str(start_index)].value)		
			if(end_index!=self.column_length):
				end_index = end_index + 1
			end_time = str(self.ws['I' + str(end_index)].value)
			interval = (datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')-datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')).seconds/60
			driving_area_time_interval[i] = interval
		return driving_area_time_interval

	def get_driving_time_sum(self, driving_area_time_interval):
		time_sum = 0
		for i in range(len(driving_area_time_interval)):
			time_sum = time_sum+driving_area_time_interval[i]
		return time_sum
		
	def get_driving_lon_lat(self, driving_area):
		driving_lon_lat = {}
		for i in range(len(driving_area)):
			length = len(driving_area[i])
			start_index = driving_area[i][0]
			end_index = driving_area[i][length-1]
			if(end_index!=self.column_length):
				end_index = end_index + 1
			lon_start = float(self.ws['D' + str(start_index)].value)
			lat_start = float(self.ws['E' + str(start_index)].value)
			lon_end = float(self.ws['D' + str(end_index)].value)
			lat_end = float(self.ws['E' + str(end_index)].value)
			driving_lon_lat[i] = [lon_start, lat_start, lon_end, lat_end]
		return driving_lon_lat
	
	def get_driving_path_length_list(self, driving_area):
		driving_length = {}
		for i in range(len(driving_area)):
			length = len(driving_area[i])				
			path_length_sum = 0
			if(length>1):
				start_index = driving_area[i][0]
				end_index = driving_area[i][length-1]
				if(end_index!=self.column_length):
					end_index = end_index+1
				valid_point_list = []
				for j in range(start_index, end_index):
					if(self.__is_valid_data(j)):
						valid_point_list.append(j)
				for j in range(len(valid_point_list)):
					if(j>0):
						start_lon = float(self.ws['D' + str(valid_point_list[j-1])].value)
						start_lat = float(self.ws['E' + str(valid_point_list[j-1])].value)
						end_lon = float(self.ws['D' + str(valid_point_list[j])].value)
						end_lat = float(self.ws['E' + str(valid_point_list[j])].value)
						dist = calc_distance(start_lon, start_lat, end_lon, end_lat)
						path_length_sum = path_length_sum+dist
				
			driving_length[i] = path_length_sum
		return 	driving_length
		
	def get_driving_length_sum(self, driving_length):
		path_sum = 0
		for i in range(len(driving_length)):
			path_sum = path_sum+driving_length[i]
		return path_sum
	
	def get_speed_variance(self, driving_area, center_value):
		speed_variance = {}
		for i in range(len(driving_area)):
			j = driving_area[i][0]
			speed_sum = 0
			count = 0
			while(j<=driving_area[i][1]):
				if(self.__is_valid_data(j)):
					sub_speed_center = int(self.ws['F' + str(j)].value) - center_value
					speed_sum = speed_sum + sub_speed_center*sub_speed_center
					count = count + 1
				j = j+1
			speed_variance[i] = math.sqrt(speed_sum/count)
		return speed_variance

	def __get_reasonable_speed_points(self, driving_area, speed_low, speed_high):
		driving_reasonable_speed_points = {}
		for i in range(len(driving_area)):
			length = len(driving_area[i])
			start_index = driving_area[i][0]
			end_index = driving_area[i][length-1]
			reasonable_speed_list = []
			for j in range(start_index, end_index+1):
				current_speed = int(self.ws['F' + str(j)].value)
				if(current_speed>=speed_low and current_speed<=speed_high):
					reasonable_speed_list.append(j)
			driving_reasonable_speed_points[i] = reasonable_speed_list
		return 	driving_reasonable_speed_points
	
	def __get_reasonable_speed_area(self, driving_area, speed_low, speed_high):
		driving_reasonable_speed_points = self.__get_reasonable_speed_points(driving_area, speed_low, speed_high)
		driving_reasonable_speed_area = {}
		j = 0
		for i in range(len(driving_reasonable_speed_points)):
			reasonable_speed_point_list = driving_reasonable_speed_points[i]
			continues_point_list = []
			for point in reasonable_speed_point_list:
				if(len(continues_point_list)==0):
					continues_point_list.append(point)
				elif(point-1==continues_point_list[len(continues_point_list)-1]):
					continues_point_list.append(point)
				else:
					driving_reasonable_speed_area[j] = continues_point_list
					j = j+1
					continues_point_list = []
					continues_point_list.append(point)
		return driving_reasonable_speed_area	
		
	def get_reasonable_speed_time_interval(self, driving_area, speed_low, speed_high):
		driving_reasonable_speed_area = self.__get_reasonable_speed_area(driving_area, speed_low, speed_high)
		self.__log_to_file('reasonable_speed_area.txt', driving_reasonable_speed_area)
		time_sum = 0
		for i in range(len(driving_reasonable_speed_area)):
			if(len(driving_reasonable_speed_area[i])>1):
				length = len(driving_reasonable_speed_area[i])
				start_index = driving_reasonable_speed_area[i][0]
				end_index = driving_reasonable_speed_area[i][length-1]
				if(end_index!=self.column_length):
					end_index = end_index+1
				start_time = datetime.strptime(str(self.ws['I' + str(start_index)].value), '%Y-%m-%d %H:%M:%S')
				end_time = datetime.strptime(str(self.ws['I' + str(end_index)].value), '%Y-%m-%d %H:%M:%S')
				time_sum = time_sum + (end_time-start_time).seconds/60
		return time_sum

	def __get_acceleration(self, driving_area):
		acc_for_each_area = {}
		for i in range(len(driving_area)):
			length = len(driving_area[i])				
			start_index = driving_area[i][0]
			end_index = driving_area[i][length-1]
			if(end_index!=self.column_length):
				end_index = end_index+1
			acc_list = []
			for j in range(start_index, end_index):
				current_speed = int(self.ws['F' + str(j)].value)
				current_time = str(self.ws['I' + str(j)].value)
				if(j==start_index and start_index==1):
					acc = 0
				elif(start_index>1):
					pre_speed = int(self.ws['F' + str(j-1)].value)
					pre_time = str(self.ws['I' + str(j-1)].value)
					interval = (datetime.strptime(current_time, '%Y-%m-%d %H:%M:%S')-datetime.strptime(pre_time, '%Y-%m-%d %H:%M:%S')).seconds
					if(interval==0):
						if(len(acc_list)==0):
							acc = 0
						else:
							acc = acc_list[len(acc_list)-1]
					else:
						acc = (current_speed-pre_speed)/(3.6*interval)
				if(acc>=ACCE_ZERO_NEG and acc<=ACCE_ZERO_POS):
					acc = 0
				acc_list.append(acc)
			acc_for_each_area[i] = acc_list
			
		return acc_for_each_area
		
	def get_speed_change_area(self, driving_area):
		acc_for_each_area = self.__get_acceleration(driving_area)
		
		acc_area = {}
		acc_index_area = {}
		acc_index_count = 0
		
		for i in range(len(acc_for_each_area)):
			acc_area_list=[]
			acc_index_area_list=[]
			#is validate data
			j = len(acc_for_each_area[i])-1
			acc_pos_count = 0
			while(j>=0):
				if(acc_for_each_area[i][j]>0):
					j = j-1
					acc_pos_count = acc_pos_count+1
					if(acc_pos_count>=2):
						print('data invalid{0}:\n{1}'.format(i, acc_for_each_area[i]))
						break
				elif(acc_for_each_area[i][j]==0):
					j = j-1
				else:
					break
			if(acc_pos_count<2):
				acc_for_each_area[i].insert(0, 0)
				acc_for_each_area[i].append(0)
				first_zero = False
				last_zero = False
				for k in range(len(acc_for_each_area[i])):
					if(k==0):
						if(acc_for_each_area[i][k]==0 and acc_for_each_area[i][k+1]!=0):
							first_zero = True
							last_zero = False
							acc_area_list.append(acc_for_each_area[i][0])
							if(driving_area[i][0]>1):
								acc_index_area_list.append(driving_area[i][0]-1)
							else:
								acc_index_area_list.append(0)
					else:	
						if(k<len(acc_for_each_area[i])-1 and acc_for_each_area[i][k]==0 and acc_for_each_area[i][k+1]!=0 and not first_zero):
							first_zero = True
							last_zero = False
							acc_area_list.append(acc_for_each_area[i][k])
							acc_index_area_list.append(driving_area[i][0]+k)
						elif(acc_for_each_area[i][k]==0 and acc_for_each_area[i][k-1]!=0 and not last_zero and first_zero):
							first_zero = False
							last_zero = True
							acc_area_list.append(acc_for_each_area[i][k])
							acc_index_area_list.append(driving_area[i][0]+k)
							acc_area[acc_index_count] = acc_area_list
							acc_index_area[acc_index_count] = acc_index_area_list
							acc_index_count = acc_index_count+1
							#the zero acc should be next area
							acc_area_list = []
							acc_index_area_list = []
							if(k<len(acc_for_each_area[i])-1 and acc_for_each_area[i][k+1]!=0):
								first_zero = True
								last_zero = False
								acc_area_list.append(acc_for_each_area[i][k])
								acc_index_area_list.append(driving_area[i][0]+k)						
						else:
							if(first_zero and not last_zero):
								acc_area_list.append(acc_for_each_area[i][k])
								acc_index_area_list.append(driving_area[i][0]+k)
		return acc_area, acc_index_area
	
	def get_speed_increase_area(self, acc_area, acc_index_area):
		acc_area_speed_inc = {}
		index_area_speed_inc = {}
		count  = 0
			
		for i in range(len(acc_area)):		
			acc_positive_count = 0
			for j in range(len(acc_area[i])):
				if(acc_area[i][j]>=0):
					acc_positive_count = acc_positive_count+1
				else:
					break
			if(acc_positive_count==len(acc_area[i])-1):
				acc_area_speed_inc[count] = acc_area[i]
				index_area_speed_inc[count] = acc_index_area[i]
				count = count+1
			else:
				speed_first = int(self.ws['F' + str(acc_index_area[i][0])].value)
				speed_last = int(self.ws['F' + str(acc_index_area[i][len(acc_index_area[i])-1])].value)
				acc_firxt = acc_area[i][1]
				acc_last = acc_area[i][len(acc_area[i])-2]
				if(acc_firxt>0 and acc_last<0 and speed_last>speed_first):
					acc_area_speed_inc[count] = acc_area[i]
					index_area_speed_inc[count] = acc_index_area[i]
					count = count+1
				elif(acc_firxt<0 and acc_last>0  and speed_last>speed_first):
					acc_area_speed_inc[count] = acc_area[i]
					index_area_speed_inc[count] = acc_index_area[i]
					count = count+1
		return acc_area_speed_inc, index_area_speed_inc
		

	def get_speed_decrease_area(self, acc_area, acc_index_area):
		acc_area_speed_dec = {}
		index_area_speed_dec = {}
		count  = 0
			
		for i in range(len(acc_area)):		
			acc_negative_count = 0
			for j in range(len(acc_area[i])):
				if(acc_area[i][j]<=0):
					acc_negative_count = acc_negative_count+1
				else:
					break
			if(acc_negative_count==len(acc_area[i])-1):
				acc_area_speed_dec[count] = acc_area[i]
				index_area_speed_dec[count] = acc_index_area[i]
				count = count+1
			else:
				speed_first = int(self.ws['F' + str(acc_index_area[i][0])].value)
				speed_last = int(self.ws['F' + str(acc_index_area[i][len(acc_index_area[i])-1])].value)
				acc_firxt = acc_area[i][1]
				acc_last = acc_area[i][len(acc_area[i])-2]
				if(acc_firxt>0 and acc_last<0 and speed_last<speed_first):
					acc_area_speed_dec[count] = acc_area[i]
					index_area_speed_dec[count] = acc_index_area[i]
					count = count+1
				elif(acc_firxt<0 and acc_last>0  and speed_last<speed_first):
					acc_area_speed_dec[count] = acc_area[i]
					index_area_speed_dec[count] = acc_index_area[i]
					count = count+1
		return acc_area_speed_dec, index_area_speed_dec
		
	def get_speed_inc_dec_start_time(self, index_area):
		time_start = {}
		for i in range(len(index_area)):
			start_index = index_area[i][0]
			start_time = str(self.ws['I' + str(start_index)].value)
			time_start[i] = start_time
		return time_start
		
	def get_speed_inc_dec_end_time(self, index_area):
		time_end = {}
		for i in range(len(index_area)):
			end_index = index_area[i][len(index_area[i])-1]
			end_time = str(self.ws['I' + str(end_index)].value)
			time_end[i] = end_time
		return time_end
	
	def get_speed_inc_dec_time_interval(self, index_area):
		time_interval = {}
		for i in range(len(index_area)):
			start_index = index_area[i][0]
			start_time = datetime.strptime(str(self.ws['I' + str(start_index)].value), '%Y-%m-%d %H:%M:%S')

			end_index = index_area[i][len(index_area[i])-1]
			end_time = datetime.strptime(str(self.ws['I' + str(end_index)].value), '%Y-%m-%d %H:%M:%S')
			time_interval[i] = (end_time-start_time).seconds
		return time_interval
	
	def get_speed_inc_dec_speed(self, index_area):
		speed_dict = {}
		for i in range(len(index_area)):
			start_index = index_area[i][0]
			start_speed = int(self.ws['F' + str(start_index)].value)

			end_index = index_area[i][len(index_area[i])-1]
			end_speed = int(self.ws['F' + str(end_index)].value)

			speed_vari = end_speed-start_speed
			speed_dict[i] = speed_vari
		return speed_dict
	
	def get_speed_inc_dec_acc(self, index_area):
		acc_dict = {}
		for i in range(len(index_area)):
			start_index = index_area[i][0]
			start_time = datetime.strptime(str(self.ws['I' + str(start_index)].value), '%Y-%m-%d %H:%M:%S')
			start_speed = int(self.ws['F' + str(start_index)].value)

			end_index = index_area[i][len(index_area[i])-1]
			end_time = datetime.strptime(str(self.ws['I' + str(end_index)].value), '%Y-%m-%d %H:%M:%S')
			end_speed = int(self.ws['F' + str(end_index)].value)

			time_interval = (end_time-start_time).seconds
			acc = (end_speed-start_speed)/(3.6*time_interval)
			acc_dict[i] = acc
		return acc_dict
			
		